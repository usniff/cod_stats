import openpyxl
import os

LIST_OF_PEOPLE = ['yousef','mitesh','brian','justin','andre','chris','jakov','taha','mike','alex','david']
LIST_OF_GAMEMODES = ['hardpoint','ctf','snd']
STATS_PATH = r'C:\Users\Elbed\Desktop\Python\COD\stats.xlsx'

wb = openpyxl.load_workbook(STATS_PATH)

def update_overall(person_loc):
    overall_sheet = wb['Overall']
    row = person_loc
    update_this_score_cell = overall_sheet.cell(row = row, column =2)
    update_this_kd_cell = overall_sheet.cell(row = row, column = 2)
    
    #update score
    hp_score = wb['Hardpoint'].cell(row = row, column = 3).value
    ctf_score = wb['CTF'].cell(row = row, column = 3).value
    snd_score = wb['SND'].cell(row = row, column = 3).value
    
    total_score = hp_score + ctf_score + snd_score
    update_this_score_cell.value = total_score
    
    # #update K/D
    # hardpoint_kd = wb['Hardpoint'].cell(row = row, column = 6).value
    # ctf_kd = wb['CTF'].cell(row = row, column = 6).value
    # snd_kd = wb['SND'].cell(row = row, column = 6).value

    # total_kd = hardpoint_kd + ctf_kd + snd_kd
    # average_kd = total_kd / 3.0
    # update_this_kd_cell.value = average_kd

def excel_editor(person,game_mode):
    for i in range(len(LIST_OF_PEOPLE)):
        if person == LIST_OF_PEOPLE[i]:
            person_loc = i+2
            break
        
    score = int(input('SCORE:\n'))
    kills = int(input('KILLS:\n'))
    deaths = int(input('DEATHS\n'))
         
        
    if game_mode.lower() == 'hardpoint':
        hpcaps = int(input('HARDPOINT CAPTURES:\n'))
        hpdefends = int(input('HARDPOINT DEFENDS:\n'))
        sheet = wb['Hardpoint']            
        row = person_loc
        
        #update matches
        match_cell = sheet.cell(row = row, column = 2)
        current_matches = match_cell.value
        match_cell.value = current_matches + 1   

        #update score
        score_cell = sheet.cell(row = row, column = 3)
        current_score = score_cell.value
        score_cell.value = current_score + score

        #update kills
        kills_cell = sheet.cell(row = row, column = 4)
        current_kills = kills_cell.value
        kills_cell.value = current_kills + kills

        #update deaths
        deaths_cell = sheet.cell(row = row, column = 5)
        current_deaths = deaths_cell.value
        deaths_cell.value = current_deaths + deaths

        #update captures
        captures_cell = sheet.cell(row = row, column = 7)
        current_captures = captures_cell.value
        captures_cell.value = current_captures + hpcaps

        #update defends
        defends_cell = sheet.cell(row = row, column = 8)        
        current_defends = defends_cell.value
        defends_cell.value = current_defends + hpdefends
        
    elif game_mode.lower() == 'ctf':
        ctfcaps = int(input('CTF CAPTURES:\n'))
        returns = int(input('CTF RETURNS:\n'))
        sheet = wb['CTF']            
        row = person_loc
        
        #update matches
        match_cell = sheet.cell(row = row, column = 2)
        current_matches = match_cell.value
        match_cell.value = current_matches + 1   

        #update score
        score_cell = sheet.cell(row = row, column = 3)
        current_score = score_cell.value
        score_cell.value = current_score + score

        #update kills
        kills_cell = sheet.cell(row = row, column = 4)
        current_kills = kills_cell.value
        kills_cell.value = current_kills + kills

        #update deaths
        deaths_cell = sheet.cell(row = row, column = 5)
        current_deaths = deaths_cell.value
        deaths_cell.value = current_deaths + deaths

        #update captures
        captures_cell = sheet.cell(row = row, column = 7)
        current_captures = captures_cell.value
        captures_cell.value = current_captures + ctfcaps
        
        #update returns
        returns_cell = sheet.cell(row = row, column = 8)    
        current_returns = returns_cell.value
        returns_cell.value = current_returns + returns
        
    else:
        plants = int(input('PLANTS:\n'))
        defuses = int(input('DEFUSES:\n'))
        sheet = wb['SND']            
        row = person_loc
        
        #update matches
        match_cell = sheet.cell(row = row, column = 2)
        current_matches = match_cell.value
        match_cell.value = current_matches + 1   

        #update score
        score_cell = sheet.cell(row = row, column = 3)
        current_score = score_cell.value
        score_cell.value = current_score + score

        #update kills
        kills_cell = sheet.cell(row = row, column = 4)
        current_kills = kills_cell.value
        kills_cell.value = current_kills + kills

        #update deaths
        deaths_cell = sheet.cell(row = row, column = 5)
        current_deaths = deaths_cell.value
        deaths_cell.value = current_deaths + deaths

        #update plants
        plants_cell = sheet.cell(row = row, column = 7)
        current_plants = plants_cell.value
        plants_cell.value = current_plants + plants

        #update defuses
        defuses_cell = sheet.cell(row = row, column = 8)    
        current_defuses = defuses_cell.value
        defuses_cell.value = current_defuses + defuses
  
    update_overall(person_loc)
    wb.save(STATS_PATH)    

game_mode = input('Which game mode are you entering?\n')
if game_mode.lower() not in LIST_OF_GAMEMODES:
    print('invalid game mode, try again\n')
    exit()
    
while True:     
    person = input('Who are you entering data for?\n')
    if person.lower() not in LIST_OF_PEOPLE:
        print('invalid person, try again\n')
        continue      

    excel_editor(person,game_mode)
    print(game_mode + ' for ' + person + ' has been updated successfully\n')    
    response = input('Do you have more people to input?\n Input 1 for yes, 0 to quit\n')
    if response == '0':
        exit()
    elif response == '1':
        continue
    else:
        exit()